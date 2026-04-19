#!/usr/bin/env python3
"""
GDPVal judge — scores deliverables in one of two modes.

JUDGE_MODE=rubric    (default)  Score model output against a rubric (0.0-1.0).
JUDGE_MODE=pairwise             Compare model output vs human gold output.
                                Returns {0.0, 0.5, 1.0} = lose / tie / win
                                (matches GDPVal paper methodology).

Other env vars:
  JUDGE_MODEL     model id (e.g. claude-haiku-4-5-20251001, gemini-3.1-pro-preview)
  JUDGE_PROVIDER  gemini | anthropic | ollama  (auto-inferred from model name)

Credentials (first available wins):
  GEMINI_API_KEY / ANTHROPIC_API_KEY / CLAUDE_CODE_OAUTH_TOKEN / OLLAMA_HOST
"""

from __future__ import annotations

import base64
import json
import os
import random
import subprocess
import sys
import urllib.request
from pathlib import Path

RUBRIC_PATH = Path("/tests/rubric.json")
DELIVERABLES_PATH = Path("/tests/deliverables.txt")
PROMPT_PATH = Path("/tests/prompt.txt")
SOLUTION_DIR = Path("/tests/solution")
REWARD_PATH = Path("/logs/verifier/reward.json")
PDF_CACHE_DIR = Path("/tmp/judge-pdfs")

MAX_CONTENT_CHARS = 60_000
MAX_INLINE_BYTES = 18 * 1024 * 1024
PDF_CONVERTIBLE = {".xlsx", ".xlsm", ".xls", ".docx", ".doc",
                   ".pptx", ".ppt", ".odt", ".ods"}

RUBRIC_INSTRUCTIONS = (
    "You are an expert professional grader for the GDPVal benchmark.\n"
    "Score the deliverables provided below against the rubric.\n\n"
    "Score on a scale from 0.0 to 1.0 based on the fraction of rubric points earned.\n"
    'Respond with ONLY a JSON object: {"score": <float 0-1>, "reason": "<brief>"}\n'
)

PAIRWISE_INSTRUCTIONS = (
    "You are a professional expert grader for the GDPVal benchmark.\n"
    "Compare two deliverables (Deliverable A and Deliverable B) produced in response "
    "to the task below. One was written by an AI; the other by a human expert. "
    "You do NOT know which is which — judge purely on quality.\n\n"
    "Return ONLY a JSON object with these exact keys:\n"
    '  {"winner": "A" | "B" | "tie", "reason": "<brief>"}\n'
)

PROVIDER_DEFAULTS = {
    "gemini": "gemini-3.1-pro-preview",
    "anthropic": "claude-sonnet-4-5",
    "ollama": "gpt-oss:120b-cloud",
}
MULTIMODAL_PROVIDERS = {"gemini", "anthropic"}


# ---------- Text extraction ----------

def _extract_xlsx(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"### Sheet: {sheet_name}")
        for row in ws.iter_rows(values_only=True):
            parts.append("\t".join("" if c is None else str(c) for c in row))
    return "\n".join(parts)


def _extract_pdf(path: Path) -> str:
    import pdfplumber
    with pdfplumber.open(path) as pdf:
        return "\n\n".join((page.extract_text() or "") for page in pdf.pages)


def _extract_docx(path: Path) -> str:
    from docx import Document
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs)


_TEXT_EXTRACTORS = {
    ".xlsx": _extract_xlsx, ".xlsm": _extract_xlsx, ".xls": _extract_xlsx,
    ".pdf": _extract_pdf,
    ".docx": _extract_docx,
}


def extract_text(path: Path) -> str:
    ext = path.suffix.lower()
    try:
        content = _TEXT_EXTRACTORS[ext](path) if ext in _TEXT_EXTRACTORS \
            else path.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        return f"[extraction failed for {path.name}: {e}]"
    if len(content) > MAX_CONTENT_CHARS:
        content = content[:MAX_CONTENT_CHARS] + "\n[...truncated...]"
    return content


# ---------- Attachment ----------

class Attachment:
    """A deliverable file exposed either as inline PDF bytes or extracted text."""

    def __init__(self, name: str, path: Path):
        self.name = name
        self.path = path

    def as_pdf_or_text(self) -> tuple[str, str]:
        pdf = self._convert_to_pdf()
        if pdf is not None:
            b64 = self._inline_b64(pdf)
            if b64 is not None:
                return "pdf", b64
        return "text", extract_text(self.path)

    def as_text(self) -> str:
        return extract_text(self.path)

    def _convert_to_pdf(self) -> Path | None:
        ext = self.path.suffix.lower()
        if ext == ".pdf":
            return self.path
        if ext not in PDF_CONVERTIBLE:
            return None
        PDF_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        target = PDF_CACHE_DIR / (self.path.stem + ".pdf")
        if target.exists():
            return target
        try:
            subprocess.run(
                ["libreoffice", "--headless", "--convert-to", "pdf",
                 "--outdir", str(PDF_CACHE_DIR), str(self.path)],
                check=True, capture_output=True, timeout=120,
            )
        except Exception as e:
            print(f"PDF conversion failed for {self.path.name}: {e}", file=sys.stderr)
            return None
        return target if target.exists() else None

    @staticmethod
    def _inline_b64(path: Path) -> str | None:
        if path.stat().st_size > MAX_INLINE_BYTES:
            return None
        return base64.b64encode(path.read_bytes()).decode()


# ---------- HTTP ----------

def http_post_json(url: str, body: dict, headers: dict, timeout: int = 300) -> dict:
    data = json.dumps(body).encode()
    req = urllib.request.Request(
        url, data=data,
        headers={"Content-Type": "application/json", **headers},
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read())


def _parse_json_blob(text: str) -> dict:
    start = text.find("{")
    end = text.rfind("}") + 1
    return json.loads(text[start:end])


# ---------- Judge ----------

def _infer_provider(model: str) -> str:
    if model.startswith("gemini"):
        return "gemini"
    if model.startswith("claude"):
        return "anthropic"
    return "ollama"


def _auto_detect() -> tuple[str, str]:
    if os.environ.get("GEMINI_API_KEY"):
        return "gemini", PROVIDER_DEFAULTS["gemini"]
    if os.environ.get("ANTHROPIC_API_KEY") or os.environ.get("CLAUDE_CODE_OAUTH_TOKEN"):
        return "anthropic", PROVIDER_DEFAULTS["anthropic"]
    return "ollama", PROVIDER_DEFAULTS["ollama"]


class Judge:
    """Single configurable judge supporting rubric and pairwise scoring."""

    def __init__(self):
        model = os.environ.get("JUDGE_MODEL", "")
        provider = os.environ.get("JUDGE_PROVIDER", "")

        if provider and not model:
            model = PROVIDER_DEFAULTS.get(provider, provider)
        elif model and not provider:
            provider = _infer_provider(model)
        elif not provider and not model:
            provider, model = _auto_detect()

        self.provider = provider
        self.model = model
        self.multimodal = provider in MULTIMODAL_PROVIDERS

    # -- Rubric scoring (0.0-1.0) --

    def score_rubric(self, rubric: str, attachments: list[Attachment]) -> dict:
        prompt = f"{RUBRIC_INSTRUCTIONS}\nRUBRIC:\n{rubric}\n"
        result = self._call(prompt, [("", attachments)])
        score = max(0.0, min(1.0, float(result.get("score", 0.5))))
        return {"score": score, "reason": str(result.get("reason", ""))[:500]}

    # -- Pairwise scoring (0.0 = lose, 0.5 = tie, 1.0 = win) --

    def score_pairwise(
        self,
        task_prompt: str,
        model_attachments: list[Attachment],
        gold_attachments: list[Attachment],
    ) -> dict:
        # Randomize A/B order to blind the judge against position bias.
        model_is_a = random.random() < 0.5
        if model_is_a:
            groups = [("A (model)", model_attachments), ("B (human)", gold_attachments)]
        else:
            groups = [("A (human)", gold_attachments), ("B (model)", model_attachments)]

        prompt = (
            f"{PAIRWISE_INSTRUCTIONS}\n"
            f"TASK:\n{task_prompt}\n"
        )
        # Strip the role labels for the prompt (they were for our bookkeeping).
        blind_groups = [(g[0].split()[0], g[1]) for g in groups]
        result = self._call(prompt, blind_groups)

        winner = str(result.get("winner", "tie")).strip().upper()
        if winner == "A":
            score = 1.0 if model_is_a else 0.0
        elif winner == "B":
            score = 0.0 if model_is_a else 1.0
        else:
            score = 0.5

        return {
            "score": score,
            "reason": str(result.get("reason", ""))[:500],
            "model_was": "A" if model_is_a else "B",
            "raw_winner": winner,
        }

    # -- Provider dispatch --

    def _call(self, prompt: str, groups: list[tuple[str, list[Attachment]]]) -> dict:
        dispatch = {
            "gemini": self._call_gemini,
            "anthropic": self._call_anthropic,
            "ollama": self._call_ollama,
        }
        fn = dispatch.get(self.provider)
        if fn is None:
            raise ValueError(f"Unknown provider: {self.provider}")
        return fn(prompt, groups)

    # -- Gemini --

    def _call_gemini(self, prompt: str, groups: list[tuple[str, list[Attachment]]]) -> dict:
        key = os.environ.get("GEMINI_API_KEY")
        if not key:
            raise RuntimeError("GEMINI_API_KEY not set")
        url = (
            "https://generativelanguage.googleapis.com/v1beta/models/"
            f"{self.model}:generateContent?key={key}"
        )
        parts: list[dict] = [{"text": prompt}]
        for label, atts in groups:
            if label:
                parts.append({"text": f"\n=== Deliverable {label} ==="})
            for att in atts:
                self._append_gemini(parts, att)

        data = http_post_json(url, {
            "contents": [{"parts": parts}],
            "generationConfig": {"temperature": 0.0, "maxOutputTokens": 800},
        }, {})
        return _parse_json_blob(data["candidates"][0]["content"]["parts"][0]["text"])

    def _append_gemini(self, parts: list[dict], att: Attachment):
        if self.multimodal:
            kind, payload = att.as_pdf_or_text()
            if kind == "pdf":
                parts.append({"text": f"\n--- {att.name} (attached PDF) ---"})
                parts.append({"inline_data": {"mime_type": "application/pdf", "data": payload}})
                return
        parts.append({"text": f"\n--- {att.name} ---\n{att.as_text()}"})

    # -- Anthropic --

    def _call_anthropic(self, prompt: str, groups: list[tuple[str, list[Attachment]]]) -> dict:
        headers = self._anthropic_headers()
        if headers is None:
            raise RuntimeError("No Anthropic credentials")

        content: list[dict] = [{"type": "text", "text": prompt}]
        for label, atts in groups:
            if label:
                content.append({"type": "text", "text": f"\n=== Deliverable {label} ==="})
            for att in atts:
                self._append_anthropic(content, att)

        data = http_post_json("https://api.anthropic.com/v1/messages", {
            "model": self.model, "max_tokens": 800, "temperature": 0.0,
            "messages": [{"role": "user", "content": content}],
        }, headers)
        return _parse_json_blob(data["content"][0]["text"])

    def _append_anthropic(self, content: list[dict], att: Attachment):
        if self.multimodal:
            kind, payload = att.as_pdf_or_text()
            if kind == "pdf":
                content.append({"type": "text", "text": f"\n--- {att.name} (attached PDF) ---"})
                content.append({
                    "type": "document",
                    "source": {"type": "base64", "media_type": "application/pdf", "data": payload},
                })
                return
        content.append({"type": "text", "text": f"\n--- {att.name} ---\n{att.as_text()}"})

    @staticmethod
    def _anthropic_headers() -> dict | None:
        if key := os.environ.get("ANTHROPIC_API_KEY"):
            return {"x-api-key": key, "anthropic-version": "2023-06-01"}
        if oauth := os.environ.get("CLAUDE_CODE_OAUTH_TOKEN"):
            return {
                "Authorization": f"Bearer {oauth}",
                "anthropic-version": "2023-06-01",
                "anthropic-beta": "oauth-2025-04-20",
            }
        return None

    # -- Ollama --

    def _call_ollama(self, prompt: str, groups: list[tuple[str, list[Attachment]]]) -> dict:
        host = os.environ.get("OLLAMA_HOST", "http://host.docker.internal:11434")
        body = prompt
        for label, atts in groups:
            if label:
                body += f"\n=== Deliverable {label} ===\n"
            for att in atts:
                body += f"\n--- {att.name} ---\n{att.as_text()}"

        data = http_post_json(f"{host}/api/generate", {
            "model": self.model, "prompt": body, "stream": False,
            "options": {"temperature": 0.0},
        }, {}, timeout=300)
        return _parse_json_blob(data.get("response", ""))


# ---------- Deliverable discovery & output ----------

def find_deliverable(name: str, roots: tuple[Path, ...]) -> Path | None:
    basename = Path(name).name
    for root in roots:
        exact = root / name
        if exact.is_file():
            return exact
        if root.is_dir():
            for match in root.rglob(basename):
                if match.is_file():
                    return match
    return None


def write_reward(reward: float, reason: str, judge: str = "none", extras: dict | None = None) -> None:
    REWARD_PATH.parent.mkdir(parents=True, exist_ok=True)
    REWARD_PATH.write_text(json.dumps({"reward": reward}))
    payload = {"reward": reward, "reason": reason, "judge": judge}
    if extras:
        payload.update(extras)
    (REWARD_PATH.parent / "judge.json").write_text(json.dumps(payload, indent=2))


def _collect(names: list[str], roots: tuple[Path, ...]) -> tuple[list[Attachment], list[str]]:
    attachments: list[Attachment] = []
    missing: list[str] = []
    for name in names:
        path = find_deliverable(name, roots)
        if path is None:
            missing.append(name)
        else:
            attachments.append(Attachment(name, path))
    return attachments, missing


def main() -> None:
    mode = os.environ.get("JUDGE_MODE", "rubric").lower()
    deliverables = [
        d.strip() for d in DELIVERABLES_PATH.read_text().splitlines() if d.strip()
    ]

    model_attachments, missing = _collect(
        deliverables, (Path("/app/output"), Path("/app"))
    )
    if missing:
        write_reward(0.0, f"missing model deliverables: {', '.join(missing)}")
        print(f"Missing model deliverables: {missing}. Reward: 0.0")
        return

    judge = Judge()
    print(f"Judge: mode={mode} provider={judge.provider} model={judge.model}")

    try:
        if mode == "pairwise":
            gold_attachments, gold_missing = _collect(deliverables, (SOLUTION_DIR,))
            if gold_missing:
                write_reward(0.0, f"missing gold deliverables: {', '.join(gold_missing)}")
                return
            task_prompt = PROMPT_PATH.read_text(encoding="utf-8") if PROMPT_PATH.exists() else ""
            result = judge.score_pairwise(task_prompt, model_attachments, gold_attachments)
        else:
            rubric = RUBRIC_PATH.read_text() if RUBRIC_PATH.exists() else "[]"
            result = judge.score_rubric(rubric, model_attachments)
    except Exception as e:
        print(f"Judge errored: {e}", file=sys.stderr)
        write_reward(0.5, f"judge errored: {e}", judge=judge.provider)
        return

    extras = {k: v for k, v in result.items() if k not in {"score", "reason"}}
    extras["mode"] = mode
    write_reward(
        result["score"], result["reason"],
        judge=f"{judge.provider}/{judge.model}", extras=extras,
    )
    print(f"Reward: {result['score']}. Reason: {result['reason']}")


if __name__ == "__main__":
    main()
